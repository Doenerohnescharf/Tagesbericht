#!/usr/bin/env python
# -*- coding: utf-8 -*-
import logging
import configparser
import argparse
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import sys
import os

def setup_logging(config):
    log_level_str = config['Logging'].get('level', 'INFO').upper()
    log_level = getattr(logging, log_level_str, logging.WARNING)
    log_targets = [target.strip().lower() for target in config['Logging'].get('targets', 'console').split(',')]
    log_file_path = config['Logging'].get('file_path', 'Tagesbericht.log')

    # Clear existing handlers (if any)
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)

    handlers = []
    if 'console' in log_targets:
        handlers.append(logging.StreamHandler(sys.stdout))  # Ensure output to stdout
    if 'file' in log_targets:
        handlers.append(logging.FileHandler(log_file_path))

    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=handlers
    )

def load_config(config_path='Tagesbericht.ini'):
    config = configparser.ConfigParser()
    try:
        config.read(config_path)
        logging.info(f"Configuration loaded from {config_path}")
        return config
    except configparser.Error as e:
        logging.error(f"Error reading configuration file: {e}")
        sys.exit(1)

def parse_args():
    parser = argparse.ArgumentParser(description='Generate Tagesberichte from SQLite database.')
    parser.add_argument('--date', help='Specific date to generate the report for.')
    parser.add_argument('--date-from', help='Start date for date range filter.')
    parser.add_argument('--date-to', help='End date for date range filter.')
    return parser.parse_args()

def check_database_exists(db_path):
    if not os.path.exists(db_path):
        logging.error(f"Database file not found: {db_path}")
        return False
    return True

def generate_output_filename(config, date=None, date_from=None, date_to=None):
    settings = config['Settings']
    output_path = settings.get('output_path', '.')
    prefix = settings.get('output_prefix', 'Tagesbericht')
    
    # Generate the date portion of the filename
    if date:
        date_suffix = f"_{date}"
    elif date_from and date_to:
        date_suffix = f"_{date_from}_to_{date_to}"
    else:
        date_suffix = ""
    
    filename = f"{prefix}{date_suffix}.xlsx"
    full_path = os.path.join(output_path, filename)
    
    # Ensure output directory exists
    os.makedirs(output_path, exist_ok=True)
    
    return full_path, filename

def format_date(date_string):
    if pd.isna(date_string) or date_string == '':
        return ''
    try:
        return datetime.strptime(date_string, '%Y-%m-%d').strftime('%d.%m.%Y')
    except ValueError:
        logging.warning(f"Unable to parse date: {date_string}")
        return date_string
    
def get_latest_date(db_path):
    try:
        with sqlite3.connect(db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT MAX(wz__dat) FROM el_pwz")
            latest_date = cursor.fetchone()[0]
            logging.info(f"Latest date found in database: {latest_date}")
            return latest_date
    except sqlite3.Error as e:
        logging.error(f"Error getting latest date from database: {e}")
        return None

def get_data_from_sqlite(db_path, mandant, columns, date_from=None, date_to=None):
    try:
        with sqlite3.connect(db_path) as conn:
            # Start with base query and parameters
            query = f"SELECT {', '.join(columns)} FROM el_pwz WHERE mandant = ?"
            params = [mandant]

            # Add date filters if specified
            if date_from and date_to:
                query += " AND wz__dat BETWEEN ? AND ?"
                params.extend([date_from, date_to])
            elif date_from:
                query += " AND wz__dat >= ?"
                params.append(date_from)
            elif date_to:
                query += " AND wz__dat <= ?"
                params.append(date_to)

            logging.info(f"Executing query: {query} with params: {params}")
            df = pd.read_sql_query(query, conn, params=params)
            logging.info(f"Retrieved {len(df)} rows for mandant {mandant}")
            return df
    except sqlite3.Error as e:
        logging.error(f"Database error: {e}")
        return pd.DataFrame()

def create_excel_sheet(workbook, sheet_name, df, column_names):
    sheet = workbook.create_sheet(title=sheet_name)
    sheet.freeze_panes = "A2"  # Freezes the top row
    
    fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    # Write the column headers
    for col, header in enumerate(df.columns, start=1):
        display_name = column_names.get(header, header)
        cell = sheet.cell(row=1, column=col, value=display_name)
        cell.fill = fill
        sheet.column_dimensions[cell.column_letter].width = max(len(display_name), 10)

    # Write the data
    for row, data in enumerate(df.values, start=2):
        for col, value in enumerate(data, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            
            # Adjust column width based on the length of the content in each cell
            column_letter = cell.column_letter
            current_width = sheet.column_dimensions[column_letter].width
            value_length = len(str(value)) if value else 0
            
            # Set the new width if the value is longer than the current width
            if value_length > current_width:
                sheet.column_dimensions[column_letter].width = value_length

def sqlite_to_xlsx(db_path, xlsx_path, columns, mandants, column_names, date=None, date_from=None, date_to=None):
    try:
        # Create the output directory if it doesn't exist
        output_dir = os.path.dirname(xlsx_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
            
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet

        for mandant, name in mandants.items():
            # If a specific date is provided, use it for both date_from and date_to
            if date:
                df = get_data_from_sqlite(db_path, mandant, columns, date_from=date, date_to=date)
            else:
                df = get_data_from_sqlite(db_path, mandant, columns, date_from=date_from, date_to=date_to)
            
            if not df.empty:
                # Format date columns
                for col in ['wz__dat', 'wz__geb']:
                    if col in df.columns:
                        df[col] = df[col].apply(format_date)

                create_excel_sheet(workbook, name, df, column_names)

        workbook.save(xlsx_path)
        logging.info(f"Excel file created successfully at {xlsx_path}")
    except PermissionError:
        logging.error(f"Permission denied: Unable to save file at {xlsx_path}")
    except Exception as e:
        logging.error(f"Error saving Excel file: {e}")

def main():
    args = parse_args()
    config = load_config()
    setup_logging(config)

    settings = config['Settings']
    columns = settings.get('columns', '').split(',')
    database = settings.get('database', 'out.sqlite')
    date_from = settings.get('date_from')
    date_to = settings.get('date_to')
    date = settings.get('date')

    if not check_database_exists(database):
        sys.exit(1)

    mandants = dict(config['Mandants'])
    column_names = dict(config['Column_Names'])

    # Override config settings with command-line args
    if args.date:
        date = args.date
    if args.date_from:
        date_from = args.date_from
    if args.date_to:
        date_to = args.date_to    

    if not date and not date_from and not date_to:
        # If no date is specified, use the latest date from the database
        latest_date = get_latest_date(database)
        if latest_date:
            date = latest_date
            logging.info(f"Using latest date from database: {date}")
        else:
            logging.error("No date specified and couldn't retrieve latest date from database.")
            sys.exit(1)

    # Generate output filename based on configuration and dates
    output_file, filename = generate_output_filename(config, date, date_from, date_to)
    logging.info(f"Output will be saved to: {output_file}")

    if date:
        sqlite_to_xlsx(database, output_file, columns, mandants, column_names, date=date)
    else:
        sqlite_to_xlsx(database, output_file, columns, mandants, column_names, date_from=date_from, date_to=date_to)

if __name__ == '__main__':
    main()